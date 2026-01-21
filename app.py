import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
import io 
import random
import datetime
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import altair as alt 

# --- 1. CONFIGURAÇÃO E CSS (VISUAL) ---
st.set_page_config(page_title="Pesquisador de Preços", page_icon="🔎", layout="wide")

st.markdown("""
<style>
    /* Importando Fonte Moderna */
    @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;700&display=swap');

    /* Reset Geral */
    * {font-family: 'Roboto', sans-serif;}
    .stApp {background-color: #f0f2f6 !important; color: #31333F !important;}
    
    /* Esconder elementos técnicos */
    .stException, #MainMenu, footer {display: none !important;}
    
    /* CABEÇALHO ESTILIZADO */
    .header-style {
        background: linear-gradient(90deg, #0052cc 0%, #0073e6 100%);
        padding: 20px;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 20px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    
    /* BOTÃO DE BUSCA (PULSANTE) */
    div.stButton > button:first-child {
        background: linear-gradient(45deg, #28a745, #218838);
        color: white;
        border: none;
        border-radius: 30px;
        height: 55px;
        font-size: 18px;
        font-weight: bold;
        text-transform: uppercase;
        letter-spacing: 1px;
        box-shadow: 0 4px 10px rgba(40, 167, 69, 0.3);
        transition: transform 0.2s;
    }
    div.stButton > button:first-child:hover {
        transform: scale(1.02);
    }

    /* CARDS DO PÓDIO */
    .podium-card {
        background-color: white;
        padding: 20px;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        text-align: center;
        border-top: 5px solid #ddd;
        height: 100%;
    }
    .gold {border-color: #FFD700; background: linear-gradient(to bottom, #fff, #fffbf0);}
    .silver {border-color: #C0C0C0; background: linear-gradient(to bottom, #fff, #f8f9fa);}
    .bronze {border-color: #CD7F32; background: linear-gradient(to bottom, #fff, #fff5f0);}
    
    /* TEXTOS */
    .price-tag {font-size: 24px; font-weight: bold; color: #212529;}
    .store-tag {font-size: 14px; color: #6c757d; text-transform: uppercase;}
</style>
""", unsafe_allow_html=True)

# --- 2. VARIÁVEIS DO SISTEMA ---
if 'dados' not in st.session_state: st.session_state.dados = []
if 'status_ml' not in st.session_state: st.session_state.status_ml = "Aguardando"
if 'status_amz' not in st.session_state: st.session_state.status_amz = "Aguardando"

# --- 3. LÓGICA INTELIGENTE ---
def converter_preco(texto):
    try:
        if isinstance(texto, (float, int)): return float(texto)
        limpo = str(texto).replace("R$", "").replace(" ", "").replace("\xa0", "")
        if "," in limpo and "." in limpo: limpo = limpo.replace(".", "").replace(",", ".")
        elif "," in limpo: limpo = limpo.replace(",", ".")
        return float(limpo)
    except: return 0.0

def calcular_similaridade(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def obter_headers():
    agentes = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ]
    return {"User-Agent": random.choice(agentes), "Accept-Language": "pt-BR,pt;q=0.9"}

# --- 4. ROBÔS DE BUSCA ---
def buscar_mercadolivre(produto):
    url = f"https://lista.mercadolivre.com.br/{produto.replace(' ', '-')}"
    try:
        resp = requests.get(url, headers=obter_headers(), timeout=6)
        if resp.status_code == 200: st.session_state.status_ml = "Online ✅"
        else: st.session_state.status_ml = "Instável ⚠️"
        
        soup = BeautifulSoup(resp.text, 'html.parser')
        itens = soup.find_all('div', class_='poly-card__content')
        if not itens: itens = soup.find_all('li', class_='ui-search-layout__item')
        if not itens: itens = soup.find_all('div', class_='ui-search-result__wrapper')
        
        lista = []
        for item in itens[:15]: 
            try:
                tag = item.find('h2') or item.find('a', class_='poly-component__title')
                titulo = tag.text.strip() if tag else "Sem Título"
                
                cont = item.find('div', class_='poly-price__current') or item.find('div', class_='ui-search-price__second-line')
                txt = "0"
                if cont:
                    span = cont.find('span', class_='andes-money-amount__fraction')
                    if span: txt = span.text
                
                link = item.find('a')['href']
                val = converter_preco(txt)
                
                if val > 5: 
                    lista.append({'Loja': 'Mercado Livre', 'Produto': titulo, 'Valor_Visual': f"R$ {txt}", 'Valor_Numerico': val, 'Link': link, 'Tipo': 'Auto'})
            except: continue
        return lista
    except: 
        st.session_state.status_ml = "Erro ❌"
        return []

def buscar_amazon(produto):
    url = f"https://www.amazon.com.br/s?k={produto.replace(' ', '+')}"
    try:
        resp = requests.get(url, headers=obter_headers(), timeout=6)
        if resp.status_code == 200: st.session_state.status_amz = "Online ✅"
        else: st.session_state.status_amz = "Bloqueio 🔒"

        soup = BeautifulSoup(resp.text, 'html.parser')
        itens = soup.find_all('div', {'data-component-type': 's-search-result'})
        lista = []
        for item in itens[:10]:
            try:
                titulo = item.find('h2').text.strip()
                whole = item.find('span', class_='a-price-whole')
                fraction = item.find('span', class_='a-price-fraction')
                if whole:
                    val_txt = f"{whole.text}{fraction.text if fraction else '00'}"
                    link = "https://www.amazon.com.br" + item.find('a', class_='a-link-normal')['href']
                    val = converter_preco(val_txt)
                    lista.append({'Loja': 'Amazon', 'Produto': titulo, 'Valor_Visual': f"R$ {val_txt}", 'Valor_Numerico': val, 'Link': link, 'Tipo': 'Auto'})
            except: continue
        return lista
    except: 
        st.session_state.status_amz = "Erro ❌"
        return []

def gerar_links_extras(termo):
    return [
        {'Loja': 'Magalu', 'Link': f"https://www.magazineluiza.com.br/busca/{termo.replace(' ', '+')}/"},
        {'Loja': 'Shopee', 'Link': f"https://shopee.com.br/search?keyword={termo.replace(' ', '%20')}"},
        {'Loja': 'Casas Bahia', 'Link': f"https://www.casasbahia.com.br/{termo.replace(' ', '-')}/b"},
        {'Loja': 'Kabum', 'Link': f"https://www.kabum.com.br/busca/{termo.replace(' ', '-')}"},
        {'Loja': 'Google Shopping', 'Link': f"https://www.google.com/search?q={termo.replace(' ', '+')}&tbm=shop"}
    ]

# --- 5. INTERFACE (BARRA LATERAL) ---
with st.sidebar:
    st.header("🎛️ Painel de Controle")
    produto_input = st.text_input("O que você procura?", placeholder="Ex: iPhone 15")
    
    st.markdown("### ⚙️ Filtros")
    usar_ia = st.toggle("🛡️ Remover Acessórios (IA)", value=True, help="Remove itens muito baratos automaticamente.")
    precisao = st.slider("🎯 Precisão do Nome (%)", 0, 100, 0)
    
    if st.button("🔎 PESQUISAR AGORA"):
        st.session_state.dados = []
        with st.spinner("Pesquisando preços..."):
            ml = buscar_mercadolivre(produto_input)
            amz = buscar_amazon(produto_input)
            st.session_state.dados = ml + amz

    st.markdown("---")
    st.markdown("### 📝 Adicionar Manualmente")
    with st.form("manual"):
        l = st.selectbox("Loja", ["Magalu", "Shopee", "Outra"])
        p = st.text_input("Preço", placeholder="1000,00")
        d = st.text_input("Nome", value=produto_input)
        if st.form_submit_button("➕ Adicionar"):
            v = converter_preco(p)
            if v > 0:
                st.session_state.dados.append({
                    'Loja': l, 'Produto': f"{d} (Manual)",
                    'Valor_Visual': f"R$ {v:,.2f}", 'Valor_Numerico': v,
                    'Link': '#', 'Tipo': 'Manual'
                })
                st.success("Adicionado!")

    if st.button("Limpar Tudo"):
        st.session_state.dados = []
        st.rerun()

    st.markdown("---")
    st.caption("📡 Status da Conexão:")
    st.caption(f"ML: {st.session_state.status_ml}")
    st.caption(f"Amazon: {st.session_state.status_amz}")

# --- 6. TELA PRINCIPAL ---
st.markdown("<div class='header-style'><h1>🔎 Pesquisador de Preços <br><span style='font-size:16px'>Edição Profissional</span></h1></div>", unsafe_allow_html=True)

# Atalhos de Lojas
if produto_input:
    st.markdown("##### 🌍 Pesquisa Rápida (Outras Lojas):")
    cols = st.columns(5)
    links = gerar_links_extras(produto_input)
    for i, link in enumerate(links):
        with cols[i]:
            st.link_button(f"🔎 {link['Loja']}", link['Link'], use_container_width=True)
    st.divider()

# Processamento dos Resultados
df = pd.DataFrame(st.session_state.dados)

if not df.empty:
    if produto_input:
        df['Similaridade'] = df.apply(lambda x: 1.0 if x['Tipo'] == 'Manual' else calcular_similaridade(produto_input, x['Produto']), axis=1)
        df = df[df['Similaridade'] >= (precisao / 100.0)]

    # IA de Limpeza (Remove lixo)
    if usar_ia and len(df) > 3:
        mediana = df['Valor_Numerico'].median()
        corte = mediana * 0.3 
        df = df[df['Valor_Numerico'] > corte]

    df = df.sort_values(by="Valor_Numerico")
    
    if not df.empty:
        # --- PODIUM (TOP 3) ---
        st.markdown("### 🏆 Melhores Ofertas do Momento")
        top3 = df.head(3).reset_index(drop=True)
        c1, c2, c3 = st.columns(3)
        
        if len(top3) > 0:
            with c1:
                st.markdown(f"""
                <div class="podium-card gold">
                    <h2>🥇 1º Lugar</h2>
                    <div class="price-tag">{top3.iloc[0]['Valor_Visual']}</div>
                    <div class="store-tag">{top3.iloc[0]['Loja']}</div>
                    <p style='font-size:12px'>{top3.iloc[0]['Produto'][:40]}...</p>
                </div>
                """, unsafe_allow_html=True)
        
        if len(top3) > 1:
            with c2:
                st.markdown(f"""
                <div class="podium-card silver">
                    <h2>🥈 2º Lugar</h2>
                    <div class="price-tag">{top3.iloc[1]['Valor_Visual']}</div>
                    <div class="store-tag">{top3.iloc[1]['Loja']}</div>
                    <p style='font-size:12px'>{top3.iloc[1]['Produto'][:40]}...</p>
                </div>
                """, unsafe_allow_html=True)

        if len(top3) > 2:
            with c3:
                st.markdown(f"""
                <div class="podium-card bronze">
                    <h2>🥉 3º Lugar</h2>
                    <div class="price-tag">{top3.iloc[2]['Valor_Visual']}</div>
                    <div class="store-tag">{top3.iloc[2]['Loja']}</div>
                    <p style='font-size:12px'>{top3.iloc[2]['Produto'][:40]}...</p>
                </div>
                """, unsafe_allow_html=True)
        
        st.write("")
        st.write("")

        # --- ABAS (Gráfico, Lista, Excel) ---
        tab_chart, tab_data, tab_export = st.tabs(["📊 Gráfico Comparativo", "📋 Lista Completa", "💾 Salvar Relatório"])
        
        with tab_chart:
            st.caption("Comparando os 10 menores preços encontrados.")
            df_chart = df.head(10).copy()
            df_chart['Nome_Curto'] = df_chart['Produto'].apply(lambda x: x[:35]+"...")
            df_chart['Cor'] = ['#28a745' if i==0 else '#6c757d' for i in range(len(df_chart))]
            
            chart = alt.Chart(df_chart).mark_bar(cornerRadiusEnd=4).encode(
                x=alt.X('Valor_Numerico', title='Preço (R$)', axis=alt.Axis(grid=False, labelColor='#555')),
                y=alt.Y('Nome_Curto', sort=alt.EncodingSortField(field="Valor_Numerico", order="ascending"), title=None, axis=alt.Axis(labelColor='#555')),
                color=alt.Color('Cor', scale=None, legend=None),
                tooltip=['Loja', 'Produto', 'Valor_Visual']
            ).properties(height=350)
            
            text = chart.mark_text(align='left', dx=5, fontWeight='bold', color='#333').encode(text='Valor_Visual')
            st.altair_chart(chart + text, use_container_width=True)

        with tab_data:
            st.dataframe(df[['Loja', 'Produto', 'Valor_Visual', 'Link']], hide_index=True, use_container_width=True,
                         column_config={"Link": st.column_config.LinkColumn("Ir para Loja")})

        with tab_export:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df[['Loja', 'Produto', 'Valor_Visual', 'Link']].to_excel(writer, index=False, sheet_name='Relatorio')
            
            buffer.seek(0)
            wb = load_workbook(buffer)
            ws = wb['Relatorio']
            
            # Estilização do Excel
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            for r in range(2, ws.max_row+1):
                c = ws.cell(r, 4)
                if c.value and c.value != '#':
                    c.hyperlink = str(c.value)
                    c.value = "CLIQUE AQUI"
                    c.font = Font(color="0000FF", underline="single")
            
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            
            nome_arq = f"Relatorio_Precos_{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx"
            st.download_button("📥 Baixar Planilha Excel", data=out, file_name=nome_arq)

    else:
        st.warning(f"Produtos removidos pelo filtro de IA ou Nome.")
        st.info("Tente desligar a opção 'Remover Acessórios (IA)' na barra lateral.")

elif produto_input:
    st.info("Nenhum resultado automático. Use os botões acima para pesquisar manualmente.")
else:
    st.info("👈 Digite o produto na barra lateral para começar.")